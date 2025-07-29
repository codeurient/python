from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Excel faylını yüklə
wb = load_workbook("siyahi.xlsx")
ws = wb.active

# Sütunları təyin et (sənin verdiyin üzrə)
merge_columns = {
    "H": None,  # Ad, soyad
    "I": None,  # Vəzifə
    "J": None,  # Başlanma
    "K": None   # Bitmə
}

# Məlumatların başladığı sətir (başlıqlar 4-cü sətrə qədərdir)
start_row = 4
end_row = ws.max_row

# Sıra nömrəsinə əsasən qruplaşdırma aparırıq
current_start = start_row

for row in range(start_row + 1, end_row + 2):  # +2, sonuncunu da işləmək üçün
    # A sütununda olan "sıra nömrəsi" (əgər varsa)
    current_val = ws[f"A{row}"].value

    if current_val is not None or row == end_row + 1:
        # Qrupun sonuna çatdı, indi merge edək
        current_end = row - 1

        if current_end > current_start:
            for col_letter in merge_columns.keys():
                ws.merge_cells(f"{col_letter}{current_start}:{col_letter}{current_end}")

        # Növbəti qrupa keç
        current_start = row

# Yekunda faylı qeyd et
wb.save("siyahi_merged.xlsx")
